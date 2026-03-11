import os
import uuid
import pandas as pd
import numpy as np
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import traceback
import re
from collections import OrderedDict
import warnings
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

warnings.filterwarnings('ignore')

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app, resources={r"/*": {"origins": "*"}})

@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(e):
    return jsonify({'error': 'File too large. Maximum size is 100MB.', 'success': False}), 413

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Resource not found', 'success': False}), 404

@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({'error': 'Method not allowed', 'success': False}), 405

@app.errorhandler(500)
def internal_error(e):
    return jsonify({'error': 'Internal server error', 'success': False}), 500

@app.errorhandler(Exception)
def handle_unhandled_exception(e):
    print("Unhandled Exception:", str(e))
    traceback.print_exc()
    return jsonify({'error': 'An unexpected error occurred', 'success': False}), 500

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm', 'csv'}
MAX_FILE_SIZE = 100 * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

processed_files = {}
processed_files_lock = threading.Lock()

global_stats = {
    "totalSheetsMerged": 0,
    "todaySheetsMerged": 0,
    "lastResetDate": datetime.now().strftime("%Y-%m-%d")
}
stats_lock = threading.Lock()

EXECUTOR = ThreadPoolExecutor(max_workers=min(8, (os.cpu_count() or 2) * 2))

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def smart_detect_header(df_raw):
    if df_raw.empty:
        return 0, []

    common_header_keywords = {
        'employee', 'name', 'code', 'id', 'date', 'time', 'amount',
        'total', 'qty', 'quantity', 'price', 'cost', 'description',
        'debit', 'credit', 'balance', 'remarks', 'note', 'serial',
        'sr', 'no', 'type', 'category', 'status', 'phone', 'email'
    }

    max_score = -1
    header_candidate = 0

    for i in range(min(20, len(df_raw))):
        row = df_raw.iloc[i]
        non_empty_mask = row.notna()
        non_empty = non_empty_mask.sum()
        if non_empty == 0:
            continue
        sample = row[non_empty_mask].iloc[:10]
        text_cells = sum(1 for cell in sample if isinstance(cell, str) and cell.strip())
        text_ratio = text_cells / min(10, non_empty)
        score = non_empty + (text_ratio * 5)
        if score > max_score:
            max_score = score
            header_candidate = i

    header_row = df_raw.iloc[header_candidate]
    header_texts = [str(cell).strip().lower() for cell in header_row if pd.notna(cell)]
    keyword_matches = sum(1 for text in header_texts if any(kw in text for kw in common_header_keywords))

    if keyword_matches >= 2:
        return header_candidate, header_row.tolist()
    if header_candidate + 1 < len(df_raw) and df_raw.iloc[header_candidate + 1].notna().sum() > 0:
        return header_candidate, header_row.tolist()
    for i in range(len(df_raw)):
        if df_raw.iloc[i].notna().sum() > 0:
            return i, df_raw.iloc[i].tolist()
    return 0, []

def preserve_special_characters(text):
    if pd.isna(text):
        return ''
    text = str(text).strip()
    text = re.sub(r'[^\w\s\-_\/\\\(\)\[\]\.]', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text

def build_clean_columns(header_values):
    clean_columns = []
    for idx, col_value in enumerate(header_values):
        if pd.isna(col_value) or str(col_value).strip() == '':
            clean_columns.append(f"Column_{idx+1}")
        else:
            cleaned = preserve_special_characters(col_value)
            clean_columns.append(cleaned if cleaned else f"Column_{idx+1}")
    seen = {}
    for i, col in enumerate(clean_columns):
        if col in seen:
            seen[col] += 1
            clean_columns[i] = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
    return clean_columns

def read_excel_file_advanced(file_path, filename):
    all_sheets_data = []
    try:
        excel_file = pd.ExcelFile(file_path, engine='openpyxl')
        sheet_names = excel_file.sheet_names

        for sheet_name in sheet_names:
            try:
                df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, dtype=str)
                if df_raw.empty:
                    continue
                df_raw = df_raw.dropna(how='all', axis=0).dropna(how='all', axis=1)
                if df_raw.empty:
                    continue
                df_raw = df_raw.reset_index(drop=True)

                header_row_idx, header_values = smart_detect_header(df_raw)
                clean_columns = build_clean_columns(header_values)
                data_start = header_row_idx + 1

                if data_start >= len(df_raw):
                    continue

                data_df = df_raw.iloc[data_start:].reset_index(drop=True)

                if len(data_df.columns) > len(clean_columns):
                    extra = len(data_df.columns) - len(clean_columns)
                    clean_columns.extend([f"Column_{len(clean_columns)+i+1}" for i in range(extra)])

                data_df.columns = clean_columns[:len(data_df.columns)]
                data_df = data_df.dropna(how='all', axis=0).dropna(how='all', axis=1)
                data_df = data_df.fillna('')

                if data_df.empty:
                    continue

                data_df.insert(0, 'Source_Sheet', sheet_name)
                data_df.insert(0, 'Source_File', filename)

                all_sheets_data.append({
                    'sheet_name': sheet_name,
                    'filename': filename,
                    'tables': [{
                        'data': data_df,
                        'dataframe': data_df,
                        'header_data': [list(data_df.columns)],
                        'merged_cells': [],
                        'column_ids': list(data_df.columns),
                        'filename': filename,
                        'sheet_name': sheet_name,
                        'original_header': header_values
                    }]
                })
            except Exception as e:
                print(f"Error processing sheet {sheet_name}: {str(e)[:100]}")
                continue

        excel_file.close()
        return all_sheets_data

    except Exception as e:
        print(f"Error reading Excel file {filename}: {str(e)[:100]}")
        return read_excel_file_simple(file_path, filename)

def read_excel_file_simple(file_path, filename):
    try:
        df_dict = pd.read_excel(file_path, sheet_name=None, dtype=str)
        all_sheets_data = []
        for sheet_name, sheet_df in df_dict.items():
            if sheet_df.empty:
                continue
            sheet_df = sheet_df.reset_index(drop=True).fillna('')
            sheet_df.insert(0, 'Source_Sheet', sheet_name)
            sheet_df.insert(0, 'Source_File', filename)
            columns = list(sheet_df.columns)
            all_sheets_data.append({
                'sheet_name': sheet_name,
                'filename': filename,
                'tables': [{
                    'data': sheet_df,
                    'dataframe': sheet_df,
                    'header_data': [columns],
                    'merged_cells': [],
                    'column_ids': columns,
                    'filename': filename,
                    'sheet_name': sheet_name,
                    'original_header': columns
                }]
            })
        return all_sheets_data
    except Exception as e:
        print(f"Simple read failed for {filename}: {str(e)[:100]}")
        return []

def read_csv_file_advanced(file_path, filename):
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-16-le', 'utf-16-be']
    df = None
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding, dtype=str, on_bad_lines='skip')
            break
        except Exception:
            continue
    if df is None:
        try:
            df = pd.read_csv(file_path, dtype=str, on_bad_lines='skip')
        except Exception:
            return []
    if df is None or df.empty:
        return []
    df = df.fillna('')
    clean_columns = []
    for col in df.columns:
        if pd.isna(col) or str(col).strip() == '':
            clean_columns.append(f"Column_{len(clean_columns)+1}")
        else:
            cleaned = preserve_special_characters(col)
            clean_columns.append(cleaned if cleaned else f"Column_{len(clean_columns)+1}")
    df.columns = clean_columns
    df.insert(0, 'Source_Sheet', 'CSV_Sheet')
    df.insert(0, 'Source_File', filename)
    return [{
        'sheet_name': 'CSV_Sheet',
        'filename': filename,
        'tables': [{
            'data': df,
            'dataframe': df,
            'header_data': [clean_columns],
            'merged_cells': [],
            'column_ids': clean_columns,
            'filename': filename,
            'sheet_name': 'CSV_Sheet',
            'original_header': clean_columns
        }]
    }]

def extract_file_data(file_path, filename):
    try:
        if filename.lower().endswith('.csv'):
            return read_csv_file_advanced(file_path, filename)
        else:
            return read_excel_file_advanced(file_path, filename)
    except Exception as e:
        print(f"Error extracting data from {filename}: {str(e)[:100]}")
        traceback.print_exc()
        return []

def intelligent_column_matching(all_sheets_data):
    if not all_sheets_data:
        return []
    all_columns = OrderedDict()
    column_frequency = {}
    for sheet_data in all_sheets_data:
        for table_data in sheet_data.get('tables', []):
            df = table_data.get('dataframe')
            if df is not None:
                for col in df.columns:
                    clean_col = str(col).strip().lower()
                    column_frequency[clean_col] = column_frequency.get(clean_col, 0) + 1
                    if clean_col not in all_columns:
                        all_columns[clean_col] = col
                    elif len(str(col)) > len(str(all_columns[clean_col])):
                        all_columns[clean_col] = col
    unified_columns = []
    source_cols = ['source_file', 'source_sheet']
    for source_col in source_cols:
        if source_col in all_columns:
            unified_columns.append(all_columns.pop(source_col))
    sorted_cols = sorted(all_columns.items(), key=lambda x: column_frequency.get(x[0], 0), reverse=True)
    seen_lower = {c.lower() for c in unified_columns}
    for clean_col, orig_col in sorted_cols:
        if clean_col not in seen_lower:
            unified_columns.append(orig_col)
            seen_lower.add(clean_col)
    return unified_columns

_NUMBER_RE = re.compile(r'^-?\d+\.?\d*$')

def merge_dataframes_intelligently(all_dfs, unified_columns):
    if not all_dfs:
        return pd.DataFrame()

    unified_lower_map = {str(c).strip().lower(): c for c in unified_columns}
    aligned_dfs = []

    for df in all_dfs:
        rename_map = {}
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if col_lower in unified_lower_map:
                target = unified_lower_map[col_lower]
                if col != target:
                    rename_map[col] = target
        if rename_map:
            df = df.rename(columns=rename_map)
        df = df.reindex(columns=unified_columns)
        aligned_dfs.append(df)

    consolidated_df = pd.concat(aligned_dfs, ignore_index=True, sort=False)
    consolidated_df = consolidated_df.fillna('')

    for col in consolidated_df.columns:
        if col in ('Source_File', 'Source_Sheet'):
            continue
        series = consolidated_df[col]
        non_empty = series[series != '']
        if len(non_empty) == 0:
            continue
        sample = non_empty.iloc[:min(200, len(non_empty))]
        numeric_hits = sample.apply(
            lambda v: bool(_NUMBER_RE.match(str(v).strip())) if isinstance(v, str)
                      else isinstance(v, (int, float, np.integer, np.floating))
        ).sum()
        if numeric_hits / len(sample) > 0.5:
            try:
                consolidated_df[col] = pd.to_numeric(consolidated_df[col], errors='coerce').fillna(0)
            except Exception:
                pass

    return consolidated_df

def merge_all_data(all_sheets_data):
    if not all_sheets_data:
        return pd.DataFrame(), [], {}, {}
    all_dfs = []
    all_header_data = []
    all_merged_cells = []
    sheet_info = {}
    for sheet_data in all_sheets_data:
        if not sheet_data:
            continue
        sheet_name = sheet_data['sheet_name']
        filename = sheet_data['filename']
        for table_data in sheet_data.get('tables', []):
            df = table_data.get('dataframe')
            if df is not None and not df.empty:
                all_dfs.append(df)
                all_header_data.append(table_data.get('header_data', []))
                all_merged_cells.extend(table_data.get('merged_cells', []))
                key = f"{filename} - {sheet_name}"
                if key not in sheet_info:
                    sheet_info[key] = {'filename': filename, 'sheet_name': sheet_name, 'table_count': 0, 'row_count': len(df), 'column_count': len(df.columns)}
                sheet_info[key]['table_count'] += 1
    if not all_dfs:
        return pd.DataFrame(), [], {}, {}
    try:
        unified_columns = intelligent_column_matching(all_sheets_data)
        consolidated_df = merge_dataframes_intelligently(all_dfs, unified_columns)
    except Exception as e:
        print(f"Error in intelligent merging: {str(e)[:200]}")
        traceback.print_exc()
        try:
            consolidated_df = pd.concat(all_dfs, ignore_index=True, sort=False)
            consolidated_df = consolidated_df.fillna('')
        except Exception:
            consolidated_df = pd.DataFrame()
    return consolidated_df, all_header_data, all_merged_cells, sheet_info

# Pre-built style objects — reused for every cell (no object recreation overhead)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_HEADER_BORDER = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
_CELL_BORDER = Border(left=Side(style='thin', color="E0E0E0"), right=Side(style='thin', color="E0E0E0"), top=Side(style='thin', color="E0E0E0"), bottom=Side(style='thin', color="E0E0E0"))
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
_RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

def create_output_excel(df, output_path, header_data_list, merged_cells_list):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged_Data"

        if df.empty:
            wb.save(output_path)
            return True

        columns = df.columns.tolist()

        # Write header row with styling
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _HEADER_BORDER

        # Identify numeric columns once — avoids per-cell type checking
        float_cols = set()
        int_cols = set()
        for i, col in enumerate(columns):
            if col not in ('Source_File', 'Source_Sheet'):
                if pd.api.types.is_float_dtype(df[col]):
                    float_cols.add(i)
                elif pd.api.types.is_integer_dtype(df[col]):
                    int_cols.add(i)

        numeric_cols = float_cols | int_cols

        # Convert to list of lists — much faster than iterrows
        data_values = df.values.tolist()

        for row_idx, row_data in enumerate(data_values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = _CELL_BORDER
                i = col_idx - 1
                if i in numeric_cols:
                    cell.alignment = _RIGHT_ALIGN
                    cell.number_format = '#,##0.00' if i in float_cols else '#,##0'
                else:
                    cell.alignment = _LEFT_ALIGN

        # Auto-fit column widths (sample up to 500 rows)
        sample_limit = min(500, len(data_values))
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            i = col_idx - 1
            for row_idx in range(sample_limit):
                val = data_values[row_idx][i]
                if val is not None:
                    cell_len = len(str(val))
                    if cell_len > max_length:
                        max_length = cell_len
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        ws.freeze_panes = ws['A2']
        wb.save(output_path)
        return True

    except Exception as e:
        print(f"Error creating output Excel: {str(e)[:200]}")
        traceback.print_exc()
        return False

def _process_single_file(args):
    temp_path, orig_filename = args
    try:
        print(f"Processing: {orig_filename}")
        sheets_data = extract_file_data(temp_path, orig_filename)
        return sheets_data, orig_filename, temp_path
    except Exception as e:
        print(f"Error processing {orig_filename}: {str(e)[:200]}")
        return [], orig_filename, temp_path

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/merge', methods=['POST'])
def merge_files():
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'No files uploaded', 'success': False}), 400

        files = request.files.getlist('files')

        for f in files:
            if f and f.filename and not allowed_file(f.filename):
                return jsonify({'error': f'File {f.filename} has invalid extension', 'success': False}), 400

        valid_files = [f for f in files if f and f.filename != '' and allowed_file(f.filename)]
        if not valid_files:
            return jsonify({'error': 'No valid files selected', 'success': False}), 400

        session_id = str(uuid.uuid4())

        # Save all files first (must be in main thread for Flask file objects)
        saved_files = []
        for file in valid_files:
            safe_filename = str(uuid.uuid4()) + "_" + file.filename
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
            file.save(temp_path)
            saved_files.append((temp_path, file.filename))

        # Process files in parallel
        all_sheets_data = []
        total_tables = 0
        total_rows = 0
        total_columns = 0
        sheet_names_info = {}

        futures = {EXECUTOR.submit(_process_single_file, sf): sf for sf in saved_files}

        for future in as_completed(futures):
            sheets_data, orig_filename, temp_path = future.result()

            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception:
                pass

            if not sheets_data:
                print(f"  No data found in {orig_filename}")
                continue

            for sheet_data in sheets_data:
                sheet_name = sheet_data['sheet_name']
                key = f"{orig_filename} - {sheet_name}"
                all_sheets_data.append(sheet_data)

                if key not in sheet_names_info:
                    sheet_names_info[key] = {
                        'filename': orig_filename, 'sheet_name': sheet_name,
                        'table_count': 0, 'row_count': 0, 'column_count': 0
                    }

                for table_data in sheet_data['tables']:
                    total_tables += 1
                    df = table_data.get('dataframe', pd.DataFrame())
                    r, c = len(df), len(df.columns)
                    total_rows += r
                    total_columns = max(total_columns, c)
                    sheet_names_info[key]['table_count'] += 1
                    sheet_names_info[key]['row_count'] += r
                    sheet_names_info[key]['column_count'] = max(sheet_names_info[key]['column_count'], c)

        if not all_sheets_data:
            return jsonify({'error': 'No data found in uploaded files. Please ensure files contain data and are in supported formats (.xlsx, .xls, .xlsm, .csv).', 'success': False}), 400

        print(f"Total sheets: {len(all_sheets_data)}, tables: {total_tables}")

        try:
            consolidated_df, header_data_list, merged_cells_list, sheet_info = merge_all_data(all_sheets_data)

            if consolidated_df.empty:
                return jsonify({'error': 'No data to merge after processing', 'success': False}), 400

            print(f"Merged: {consolidated_df.shape[0]} rows x {consolidated_df.shape[1]} cols")

            # Preview — first 100 rows
            preview_cols = consolidated_df.columns.tolist()
            preview_rows_df = consolidated_df.head(100)

            def safe_val(v):
                if isinstance(v, np.integer): return int(v)
                if isinstance(v, np.floating): return float(v)
                if isinstance(v, float) and np.isnan(v): return ''
                return v

            preview_data = [preview_cols]
            for row in preview_rows_df.itertuples(index=False):
                preview_data.append([safe_val(v) for v in row])

            output_filename = f"merged_{session_id}.xlsx"
            output_path = os.path.join(UPLOAD_FOLDER, output_filename)

            success = create_output_excel(consolidated_df, output_path, header_data_list, merged_cells_list)
            if not success:
                return jsonify({'error': 'Failed to create output file', 'success': False}), 500

        except Exception as e:
            print(f"Error in merge process: {str(e)[:200]}")
            traceback.print_exc()
            return jsonify({'error': f'Error merging data: {str(e)[:200]}', 'success': False}), 500

        with processed_files_lock:
            processed_files[session_id] = {
                'filename': output_filename,
                'path': output_path,
                'created_at': datetime.now().isoformat(),
                'stats': {'tables': total_tables, 'rows': len(consolidated_df), 'columns': len(consolidated_df.columns), 'files': len(valid_files)},
                'sheet_info': sheet_names_info
            }

        with stats_lock:
            today = datetime.now().strftime("%Y-%m-%d")
            if global_stats["lastResetDate"] != today:
                global_stats["todaySheetsMerged"] = 0
                global_stats["lastResetDate"] = today
            global_stats["totalSheetsMerged"] += total_tables
            global_stats["todaySheetsMerged"] += total_tables

        return jsonify({
            'success': True,
            'download_id': session_id,
            'data': {'consolidated': preview_data},
            'stats': {'tables': total_tables, 'rows': len(consolidated_df), 'columns': len(consolidated_df.columns), 'files': len(valid_files)},
            'sheet_info': sheet_names_info
        })

    except Exception as e:
        print(f"Error in merge endpoint: {str(e)[:200]}")
        traceback.print_exc()
        return jsonify({'error': str(e)[:200], 'success': False}), 500

@app.route('/download/<session_id>', methods=['GET'])
def download_file(session_id):
    try:
        with processed_files_lock:
            if session_id not in processed_files:
                return jsonify({'error': 'File not found or expired', 'success': False}), 404
            file_info = processed_files[session_id]
        file_path = file_info['path']
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found', 'success': False}), 404
        return send_file(file_path, as_attachment=True,
                         download_name=f"Merged_Excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error in download endpoint: {str(e)[:200]}")
        traceback.print_exc()
        return jsonify({'error': str(e)[:200], 'success': False}), 500

@app.route('/cleanup', methods=['POST'])
def cleanup():
    try:
        cutoff_age = 3600
        cleaned_count = 0
        now = datetime.now().timestamp()
        with processed_files_lock:
            to_delete = []
            for sid, file_info in processed_files.items():
                file_path = file_info['path']
                if os.path.exists(file_path) and now - os.path.getmtime(file_path) > cutoff_age:
                    try: os.remove(file_path)
                    except Exception: pass
                    to_delete.append(sid)
                    cleaned_count += 1
            for sid in to_delete:
                del processed_files[sid]
        for fname in os.listdir(UPLOAD_FOLDER):
            fpath = os.path.join(UPLOAD_FOLDER, fname)
            if os.path.isfile(fpath) and now - os.path.getmtime(fpath) > cutoff_age:
                try: os.remove(fpath)
                except Exception: pass
        return jsonify({'success': True, 'cleaned': cleaned_count})
    except Exception as e:
        print(f"Error in cleanup: {str(e)[:200]}")
        traceback.print_exc()
        return jsonify({'error': str(e)[:200], 'success': False}), 500

@app.route('/stats', methods=['GET'])
def get_stats():
    return jsonify(global_stats)

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat(), 'processed_files': len(processed_files)})

@app.route('/style.css')
def serve_css():
    return send_from_directory('.', 'style.css')

@app.route('/script.js')
def serve_js():
    return send_from_directory('.', 'script.js')

if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    print("=" * 70)
    print("EXCEL MULTI-FILE MERGE TOOL - OPTIMIZED VERSION")
    print("=" * 70)
    print(f"Upload folder: {os.path.abspath(UPLOAD_FOLDER)}")
    print(f"CPU cores: {os.cpu_count()}")
    print("Server running on http://0.0.0.0:10000")
    print("=" * 70)
    try:
        now = datetime.now().timestamp()
        for fname in os.listdir(UPLOAD_FOLDER):
            fpath = os.path.join(UPLOAD_FOLDER, fname)
            if os.path.isfile(fpath) and now - os.path.getmtime(fpath) > 3600:
                os.remove(fpath)
        print("Cleaned up old files")
    except Exception:
        pass
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)
